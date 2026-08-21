from __future__ import annotations

import json
import re
import os
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook

try:
    from docx import Document
except Exception:  # pragma: no cover - optional local dependency
    Document = None


BASE_DIR = Path(os.environ.get("REPORT_BASE_DIR", Path(__file__).resolve().parent / "data"))
RUN_STATUS_PATH = BASE_DIR / "last_run_status.json"


@dataclass(frozen=True)
class ModuleConfig:
    key: str
    label: str
    folder: Path
    trigger_file: Path | None = None
    excel_folder: Path | None = None
    report_patterns: tuple[str, ...] = ()


MODULES = [
    ModuleConfig(
        key="boll",
        label="BOLL 布林线",
        folder=BASE_DIR / "BOLL",
        trigger_file=BASE_DIR / "BOLL" / "BOLL.txt",
        excel_folder=BASE_DIR / "BOLL" / "EXCLE",
        report_patterns=("美股BOLL布林线监控报告-*.docx",),
    ),
    ModuleConfig(
        key="cross",
        label="CROSS 金叉",
        folder=BASE_DIR / "CROSS",
        trigger_file=BASE_DIR / "CROSS" / "CROSS.txt",
        excel_folder=BASE_DIR / "CROSS" / "EXCLE",
        report_patterns=("强势金叉监控报告_*.docx",),
    ),
    ModuleConfig(
        key="short_risk",
        label="短线风险",
        folder=BASE_DIR / "drop" / "short-risk",
        trigger_file=BASE_DIR / "drop" / "short-risk" / "RISK.txt",
        report_patterns=("美股短线反弹风险监控_*.docx",),
    ),
    ModuleConfig(
        key="drop_boll",
        label="BOLL中长期下跌",
        folder=BASE_DIR / "drop" / "drop-BOLL",
        trigger_file=BASE_DIR / "drop" / "drop-BOLL" / "BOLL-dip.txt",
        excel_folder=BASE_DIR / "drop" / "drop-BOLL" / "EXCLE",
        report_patterns=("BOLL_MACD_RSI_Downtrend_Report_*.docx",),
    ),
    ModuleConfig(
        key="drop",
        label="中长期风险",
        folder=BASE_DIR / "drop",
        trigger_file=BASE_DIR / "drop" / "drop.txt",
        report_patterns=("美股技术风险监控_*.docx",),
    ),
]


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(str(value)).date()
    except Exception:
        return None


def clean_symbol(raw: str) -> tuple[str, int | None]:
    text = str(raw).strip()
    match = re.match(r"^([A-Za-z0-9.\-]+)[（(](\d+)[）)]$", text)
    if match:
        return match.group(1).upper(), int(match.group(2))
    return text.upper(), None


def read_text_lines(path: Path) -> list[str]:
    if not path or not path.exists():
        return []
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return [line.strip() for line in path.read_text(encoding=encoding).splitlines() if line.strip()]
        except UnicodeDecodeError:
            continue
    return []


@st.cache_data(ttl=60)
def load_run_status() -> dict:
    if not RUN_STATUS_PATH.exists():
        return {}
    try:
        return json.loads(RUN_STATUS_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        return {"状态": "状态文件读取失败", "错误": str(exc)}


@st.cache_data(ttl=60)
def load_trigger_records() -> pd.DataFrame:
    rows: list[dict] = []
    for module in MODULES:
        if not module.trigger_file:
            continue
        for line in read_text_lines(module.trigger_file):
            parts = re.split(r"\t+|\s{2,}", line)
            if len(parts) < 2:
                parts = line.split()
            if len(parts) < 2:
                continue
            trigger_date = parse_date(parts[0])
            symbol, count = clean_symbol(parts[1])
            detail = " ".join(parts[2:]).strip() if len(parts) > 2 else ""
            rows.append(
                {
                    "模块": module.label,
                    "日期": trigger_date,
                    "股票": symbol,
                    "触发次数": count,
                    "信号/评分": detail,
                    "原始记录": line,
                    "来源文件": str(module.trigger_file),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["模块", "日期", "股票", "触发次数", "信号/评分", "原始记录", "来源文件"])
    df = pd.DataFrame(rows)
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    return df.sort_values(["日期", "模块", "股票"], ascending=[False, True, True]).reset_index(drop=True)


def find_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, min(ws.max_column, 12) + 1)]
        text = "|".join(str(v) for v in values if v is not None)
        if "收盘价" in text and ("交易日" in text or "日期" in text):
            return row
    return 7 if ws.max_row >= 8 else None


def read_tracking_file(path: Path, module_label: str) -> dict | None:
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb["30日跟踪"] if "30日跟踪" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if header is None:
        return None

    headers = {
        str(ws.cell(header, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(header, col).value is not None
    }
    date_col = headers.get("交易日") or headers.get("日期") or 1
    close_col = headers.get("收盘价") or 2

    rows = []
    for row in range(header + 1, ws.max_row + 1):
        trade_date = ws.cell(row, date_col).value
        close = ws.cell(row, close_col).value
        if trade_date is None or close is None:
            continue
        try:
            rows.append((parse_date(trade_date), float(close)))
        except Exception:
            continue
    if not rows:
        return None

    trigger_date, trigger_close = rows[0]
    latest_date, latest_close = rows[-1]
    if not trigger_close:
        return None
    symbol = ws["B2"].value or path.stem
    pct = latest_close / trigger_close - 1
    return {
        "模块": module_label,
        "股票": str(symbol).upper(),
        "触发日期": trigger_date,
        "最新日期": latest_date,
        "触发价": trigger_close,
        "最新价": latest_close,
        "相对触发日涨跌幅": pct,
        "记录交易日": len(rows),
        "文件": path.name,
        "文件路径": str(path),
    }


@st.cache_data(ttl=60)
def load_tracking_records() -> pd.DataFrame:
    rows: list[dict] = []
    for module in MODULES:
        if not module.excel_folder or not module.excel_folder.exists():
            continue
        for path in sorted(module.excel_folder.glob("*.xlsx")):
            try:
                item = read_tracking_file(path, module.label)
                if item:
                    rows.append(item)
            except Exception:
                continue
    if not rows:
        return pd.DataFrame(
            columns=["模块", "股票", "触发日期", "最新日期", "触发价", "最新价", "相对触发日涨跌幅", "记录交易日", "文件", "文件路径"]
        )
    df = pd.DataFrame(rows)
    df["触发日期"] = pd.to_datetime(df["触发日期"], errors="coerce")
    df["最新日期"] = pd.to_datetime(df["最新日期"], errors="coerce")
    return df.sort_values("相对触发日涨跌幅", ascending=False).reset_index(drop=True)


@st.cache_data(ttl=60)
def load_reports() -> pd.DataFrame:
    rows = []
    for module in MODULES:
        for pattern in module.report_patterns:
            for path in sorted(module.folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True):
                report_date = None
                match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
                if match:
                    report_date = parse_date(match.group(1))
                rows.append(
                    {
                        "模块": module.label,
                        "报告日期": report_date,
                        "文件名": path.name,
                        "修改时间": datetime.fromtimestamp(path.stat().st_mtime),
                        "大小KB": round(path.stat().st_size / 1024, 1),
                        "路径": str(path),
                    }
                )
    if not rows:
        return pd.DataFrame(columns=["模块", "报告日期", "文件名", "修改时间", "大小KB", "路径"])
    df = pd.DataFrame(rows)
    df["报告日期"] = pd.to_datetime(df["报告日期"], errors="coerce")
    return df.sort_values(["报告日期", "修改时间"], ascending=[False, False]).reset_index(drop=True)


def read_docx_preview(path: str, limit: int = 18) -> str:
    if Document is None:
        return "当前环境无法读取 Word 正文预览。"
    try:
        doc = Document(path)
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs[:limit]) or "报告中没有可预览的段落。"
    except Exception as exc:
        return f"读取报告预览失败：{exc}"


def filter_by_modules(df: pd.DataFrame, modules: Iterable[str]) -> pd.DataFrame:
    if df.empty:
        return df
    selected = list(modules)
    if not selected:
        return df.iloc[0:0]
    return df[df["模块"].isin(selected)].copy()


def format_pct(value: float | None) -> str:
    if pd.isna(value):
        return ""
    return f"{value * 100:.2f}%"


def render_kpis(triggers: pd.DataFrame, tracking: pd.DataFrame, reports: pd.DataFrame) -> None:
    latest_trigger_date = triggers["日期"].max() if not triggers.empty else pd.NaT
    latest_report_date = reports["报告日期"].max() if not reports.empty else pd.NaT
    nonzero = tracking[tracking["相对触发日涨跌幅"].abs() > 1e-12] if not tracking.empty else tracking
    avg_return = nonzero["相对触发日涨跌幅"].mean() if not nonzero.empty else None

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("触发记录数", f"{len(triggers):,}")
    c2.metric("跟踪股票数", f"{len(tracking):,}")
    c3.metric("剔除0后平均涨跌幅", format_pct(avg_return) if avg_return is not None else "N/A")
    c4.metric("最新报告日期", latest_report_date.strftime("%Y-%m-%d") if pd.notna(latest_report_date) else "N/A")
    st.caption(f"最新触发日期：{latest_trigger_date.strftime('%Y-%m-%d') if pd.notna(latest_trigger_date) else 'N/A'}")


def render_run_status(status: dict) -> None:
    st.subheader("云端运行状态")
    if not status:
        st.warning("还没有读取到云端运行状态文件。可能是 GitHub Actions 尚未成功运行，或还没有把 data/last_run_status.json 提交到仓库。")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("最近运行状态", status.get("状态", "未知"))
    c2.metric("结束时间", status.get("结束时间", "N/A"))
    c3.metric("成功模块", f"{status.get('成功模块数', 0)}/{status.get('模块总数', 0)}")
    c4.metric("耗时", f"{status.get('耗时秒', 0)} 秒")

    modules = status.get("模块", [])
    if modules:
        display = pd.DataFrame(modules)
        wanted = ["模块", "状态", "退出码", "触发记录数", "本次新增触发", "最新触发日期", "开始时间", "结束时间", "耗时秒"]
        cols = [col for col in wanted if col in display.columns]
        st.dataframe(display[cols], use_container_width=True, hide_index=True)
    if status.get("错误"):
        st.error(status["错误"])


def render_tracking_section(tracking: pd.DataFrame) -> None:
    st.subheader("触发后涨跌幅排名")
    if tracking.empty:
        st.info("当前没有可读取的 Excel 跟踪数据。")
        return
    display = tracking.copy()
    display["涨跌幅"] = display["相对触发日涨跌幅"].map(format_pct)
    display = display[
        ["模块", "股票", "涨跌幅", "触发日期", "最新日期", "触发价", "最新价", "记录交易日", "文件"]
    ]
    st.dataframe(display, use_container_width=True, hide_index=True)

    top = tracking.head(20).copy()
    fig = px.bar(
        top.sort_values("相对触发日涨跌幅"),
        x="相对触发日涨跌幅",
        y="股票",
        color="模块",
        orientation="h",
        title="触发后涨跌幅 Top 20",
        labels={"相对触发日涨跌幅": "涨跌幅", "股票": "股票"},
    )
    fig.update_layout(height=560, xaxis_tickformat=".0%")
    st.plotly_chart(fig, use_container_width=True)


def render_trigger_section(triggers: pd.DataFrame) -> None:
    st.subheader("触发记录")
    if triggers.empty:
        st.info("当前没有触发记录。")
        return
    date_range = [triggers["日期"].min().date(), triggers["日期"].max().date()]
    chosen = st.date_input("触发日期范围", value=date_range, min_value=date_range[0], max_value=date_range[1])
    if isinstance(chosen, tuple) and len(chosen) == 2:
        start, end = pd.to_datetime(chosen[0]), pd.to_datetime(chosen[1])
        triggers = triggers[(triggers["日期"] >= start) & (triggers["日期"] <= end)]
    st.dataframe(triggers[["日期", "模块", "股票", "触发次数", "信号/评分", "原始记录"]], use_container_width=True, hide_index=True)


def render_reports_section(reports: pd.DataFrame) -> None:
    st.subheader("Word 报告")
    if reports.empty:
        st.info("当前没有报告文件。")
        return
    st.dataframe(reports[["报告日期", "模块", "文件名", "修改时间", "大小KB"]], use_container_width=True, hide_index=True)

    options = [f"{row['模块']} | {row['文件名']}" for _, row in reports.iterrows()]
    selected = st.selectbox("预览/下载报告", options)
    idx = options.index(selected)
    path = reports.iloc[idx]["路径"]
    with st.expander("报告正文预览", expanded=False):
        st.text(read_docx_preview(path))
    with open(path, "rb") as handle:
        st.download_button("下载这个 Word 报告", handle, file_name=Path(path).name)


def render_single_module_page(module_label: str, triggers: pd.DataFrame, tracking: pd.DataFrame, reports: pd.DataFrame) -> None:
    st.header(module_label)
    module_triggers = filter_by_modules(triggers, [module_label])
    module_tracking = filter_by_modules(tracking, [module_label])
    module_reports = filter_by_modules(reports, [module_label])
    render_kpis(module_triggers, module_tracking, module_reports)
    render_trigger_section(module_triggers)
    render_tracking_section(module_tracking)
    with st.expander("查看该模块 Word 报告", expanded=False):
        render_reports_section(module_reports)


def main() -> None:
    st.set_page_config(page_title="股票监控仪表盘", page_icon="📈", layout="wide")
    st.title("股票技术监控仪表盘")
    st.caption("云端版：读取项目 data 目录中的 TXT、Word 和 Excel 结果。")

    triggers = load_trigger_records()
    tracking = load_tracking_records()
    reports = load_reports()
    run_status = load_run_status()

    module_labels = [m.label for m in MODULES]
    selected_modules = st.sidebar.multiselect("模块筛选", module_labels, default=module_labels)
    page = st.sidebar.radio(
        "页面",
        ["总览", "BOLL", "CROSS", "短线风险", "中长期风险", "BOLL中长期下跌", "报告中心", "数据诊断"],
    )
    st.sidebar.caption(f"数据目录：{BASE_DIR}")
    if run_status:
        st.sidebar.caption(f"最近云端运行：{run_status.get('结束时间', 'N/A')} | {run_status.get('状态', '未知')}")
    if st.sidebar.button("刷新数据"):
        st.cache_data.clear()
        st.rerun()

    filtered_triggers = filter_by_modules(triggers, selected_modules)
    filtered_tracking = filter_by_modules(tracking, selected_modules)
    filtered_reports = filter_by_modules(reports, selected_modules)

    if page == "总览":
        render_run_status(run_status)
        render_kpis(filtered_triggers, filtered_tracking, filtered_reports)
        c1, c2 = st.columns([1.15, 1])
        with c1:
            render_tracking_section(filtered_tracking)
        with c2:
            st.subheader("最新触发")
            if filtered_triggers.empty:
                st.info("暂无触发记录。")
            else:
                st.dataframe(filtered_triggers.head(25)[["日期", "模块", "股票", "信号/评分"]], use_container_width=True, hide_index=True)
            st.subheader("最新报告")
            if filtered_reports.empty:
                st.info("暂无报告。")
            else:
                st.dataframe(filtered_reports.head(12)[["报告日期", "模块", "文件名"]], use_container_width=True, hide_index=True)

    elif page == "BOLL":
        modules = ["BOLL 布林线"]
        render_kpis(filter_by_modules(triggers, modules), filter_by_modules(tracking, modules), filter_by_modules(reports, modules))
        render_tracking_section(filter_by_modules(tracking, modules))
        render_trigger_section(filter_by_modules(triggers, modules))

    elif page == "CROSS":
        modules = ["CROSS 金叉"]
        render_kpis(filter_by_modules(triggers, modules), filter_by_modules(tracking, modules), filter_by_modules(reports, modules))
        render_tracking_section(filter_by_modules(tracking, modules))
        render_trigger_section(filter_by_modules(triggers, modules))

    elif page == "短线风险":
        render_single_module_page("短线风险", triggers, tracking, reports)

    elif page == "中长期风险":
        render_single_module_page("中长期风险", triggers, tracking, reports)

    elif page == "BOLL中长期下跌":
        render_single_module_page("BOLL中长期下跌", triggers, tracking, reports)

    elif page == "报告中心":
        render_reports_section(filtered_reports)

    elif page == "数据诊断":
        st.subheader("数据源状态")
        st.write(
            {
                "触发记录": len(triggers),
                "Excel跟踪记录": len(tracking),
                "Word报告": len(reports),
                "基础目录存在": BASE_DIR.exists(),
                "运行状态文件存在": RUN_STATUS_PATH.exists(),
            }
        )
        render_run_status(run_status)
        st.dataframe(pd.DataFrame([m.__dict__ for m in MODULES]), use_container_width=True)


if __name__ == "__main__":
    main()
